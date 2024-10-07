__author__ = 'Lukas Horst'

import csv

import openpyxl
import pandas as pd
from openpyxl.styles import Font

from violation_checker import ViolationChecker


def get_points(file_path: str):
    """Returns the total points"""
    data = pd.read_excel(file_path, usecols='A, C')
    total_points = 0
    for i in range(len(data)):
        if data.iat[i, 0] == 'Summe':
            break
        value = data.iat[i, 1]
        if pd.notna(value) and type(value) != str:
            total_points += value
    return max(0, total_points)


def update_style_deduction(file_path: str, violation_checker: ViolationChecker, student_name: str):
    """Function to update the deduction for style violations"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Sheet1']
    # Updating the name
    ws[f'A1'].value = student_name
    rows = ws.iter_rows(min_row=1, max_row=75, min_col=1, max_col=1)
    for i, row in enumerate(rows):
        cell = row[0]
        if cell.value is not None:
            # Updating the deduction for the author variable
            if '__author__' in cell.value:
                ws[f'C{i+1}'].value = -violation_checker.count_deduction(3)
                ws[f'C{i + 1}'].font = Font(color='FF0000')
            # All deductions except the author variable and docstrings
            elif 'alle o.g. Fehler' in cell.value:
                deduction = 0
                for j in range(1, 10):
                    if j == 3 or j == 5:
                        continue
                    deduction -= violation_checker.count_deduction(j)
                ws[f'C{i+1}'].value = deduction
                ws[f'C{i+1}'].font = Font(color='FF0000')
            # Deduction for docstrings
            elif 'Abzug bei mangelden' in cell.value:
                ws[f'C{i+1}'].value = -violation_checker.count_deduction(5)
                ws[f'C{i + 1}'].font = Font(color='FF0000')
            # Updating the function for the total points
            elif 'Summe' in cell.value:
                ws[f'C{i+1}'] = f'=MAX(0, SUM(C1:C{i}))'
                break
    wb.save(file_path)
    wb.close()


def read_csv_file(file_path: str):
    """Function to read a csv file and returns a list with each row in a dic"""
    with open(file_path, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        rows = []
        for row in reader:
            rows.append(row)
    return rows


def write_csv_file(file_path: str, data: list[dict[str, str]]):
    """Function to (over)write a csv file with the given data"""
    with open(file_path, mode='w', newline='', encoding='utf-8') as file:
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)


def update_rating(overall_rating_path: str, student_rating_path: str, student_name: str):
    """Function to update the points of the given student"""
    csv_data = read_csv_file(overall_rating_path)
    for row in csv_data:
        if row['Vollständiger Name'] == student_name:
            points = str(get_points(student_rating_path)).replace('.', ',')
            row['Bewertung'] = points
            write_csv_file(overall_rating_path, csv_data)
            return
