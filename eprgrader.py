"""
Tools for streamlining the EPR/GPR grading process.

Unpacking files downloaded from Moodle, unpacking the archives within,
running pylint on all the python files, and distributing the grading
table to all the directories.
Later on, collecting all the stylecheck files and grading tables into
neat little archives for upload.
"""

__author__ = "Adrian Welcker"
__credits__ = "Adjustments from Lukas Horst"

import argparse
import contextlib
import copy
import csv
import io
import itertools
import os
import pathlib
import platform
import shutil
import sys

import openpyxl
import pandas as pd
import unicodedata
import zipfile
import re

from openpyxl.styles import Font
from datetime import datetime

from pylint.lint import Run as RunPylint
import pycodestyle

from violation_checker import ViolationChecker

PYLINT_ARGS = [
    '--exit-zero',  # always exit with code 0, even when problems are found
    '--load-plugins=eprcheck_2019',  # load plugin for checking __author__ variable
    '--persistent=n',  # don't store results on disk
    '--score=n',  # don't output a score
    '--use-pairs=<y_or_n>',  # lint_files will set this according to command line settings
    '--disable=all',  # disable all checks for now, individual ones enabled below
    # C2100: missing-author-variable
    # C2101: malformed-author-variable
    # C2102: incorrectly-assigned-author-variable
    # C0102: blacklisted-name
    # C0103: invalid-name [e.g. variable not snake_case]
    # C0112: empty-docstring
    # C0114: missing-module-docstring
    # C0115: missing-class-docstring
    # C0116: missing-function-docstring
    # C0121: singleton-comparison
    # C0144: non-ascii-name
    "--enable=C2100,C2101,C2102,C0102,C0103,C0114,C0115,C0116,C0112,C0121,C0144" +
    # C0321: multiple-statements
    # C0325: superfluous-parens
    # C0410: multiple-imports
    # C0411: wrong-import-order
    # C0412: ungrouped-imports
    # C0413: wrong-import-position
    "C0321,C0325,C0410,C0411,C0412,C0413," +
    # E0001: syntax-error
    # E0102: function-redefined
    # E0211: no-method-argument [when it should at least have self]
    "E0001,E0102,E0211," +
    # W0104: pointless-statement
    # W0201: attribute-defined-outside-init
    # W0231: super-init-not-called
    # W0232: no-init
    # W0301: unnecessary-semicolon
    # W0311: bad-indentation
    # W0401: wildcard-import
    # W0404: reimported
    # W0603: global-statement
    # W0622: redefined-builtin
    # W0702: bare-except
    # W0705: duplicate-except
    # W0706: try-except-raise
    "W0104,W0201,W0231,W0232,W0301,W0311,W0401,W0404,W0603,W0622,W0702,W0705,W0706"
]

PYCODESTYLE_SELECT = [
    # E117: over-indented
    'E117',
    # E201: whitespace after '('
    'E201',
    # E202: whitespace before ')'
    'E202',
    # E203: whitespace before ':'
    'E203',
    # E211: whitespace before '('
    'E211',
    # E221: multiple spaces before operator
    'E221',
    # E222: multiple spaces after operator
    'E222',
    # E223: tab before operator
    'E223',
    # E224: tab after operator
    'E224',
    # E225: missing whitespace around operator
    'E225',
    # E231: missing whitespace after ',', ';', or ':'
    'E231',
    # E251: unexpected spaces around kwarg '='
    'E251',
    # E261: at least two spaces before inline comment
    'E261',
    # E262: inline comment should start with '# '
    # E265: block comment should start with '#'
    'E262', 'E265',
    'E271', 'E272', 'E273', 'E274', 'E275',
    # E302: expected 2 blank lines
    'E302',
    # E501: line-too-long
    # E502: backslash redundant between brackets
    'E501', 'E502',
    # E713: negative membership test should use 'not in'
    'E713',
    # E714: negative identitiy test should use 'is not'
    'E714',
    # E721: use 'isinstance' instead of comparing types
    'E721',
]
tmp_storage = {}
violations_checkers = {}


@contextlib.contextmanager
def pylint_context(stdout, workdir):
    """Temporarily change stdout and working directory."""
    sys.stdout = stdout
    tmp_storage['argv'] = sys.argv
    tmp_storage['workdir'] = os.getcwd()
    tmp_storage['path'] = copy.copy(sys.path)
    sys.path.append(str(pathlib.Path(__file__).parent.absolute()))
    os.chdir(workdir)
    yield
    os.chdir(tmp_storage['workdir'])
    sys.argv = tmp_storage['argv']
    sys.path = tmp_storage['path']
    sys.stdout = sys.__stdout__


def lint_files(folders, author_pairs, deduction: bool):
    """Run pylint and pycodestyle on all Python files anywhere within `folders'."""
    count = 0
    total = len(folders)
    style = pycodestyle.StyleGuide(select=PYCODESTYLE_SELECT, show_source=True)
    PYLINT_ARGS[4] = '--use-pairs=y' if author_pairs else '--use-pairs=n'
    for folder in folders:
        count += 1
        print(f" ({str(count).rjust(len(str(total)))}/{total}) Checking {folder.name}")
        pythons = list(map(pathlib.Path.resolve,
                            filter(lambda p: "__MACOSX" not in p.parts and ".venv" not in p.parts,
                                    folder.glob('**/*.py'))))
        if not pythons:
            continue
        pycount = 0
        pytotal = len(pythons) * 2
        lintcache = io.StringIO()
        for file in pythons:
            pycount += 1
            print(
                f"  ({str(pycount).rjust(len(str(pytotal)))}/{pytotal}) Running pylint for {file.name}")
            with pylint_context(lintcache, folder):
                try:
                    RunPylint(PYLINT_ARGS + [str(file)])
                except SystemExit as e:
                    if e.code:
                        print(f"  [Pylint attempted to exit with code {e.code}]", file=sys.stderr)
                        raise RuntimeError from e
                pycount += 1
                print(
                    f"  ({str(pycount).rjust(len(str(pytotal)))}/{pytotal}) Running pycodestyle for {file.name}",
                    file=sys.__stdout__)
                print('\n')
                result = style.check_files([file])
                if result.total_errors > 0:
                    print('\n')
        with open(folder / 'stylecheck.txt', 'w', encoding='utf-8') as outfile:
            if lintcache.tell() > 0:
                style_check = remove_unnecessary_violations(lintcache.getvalue())
            else:
                style_check = ""
            violation_checker = ViolationChecker(style_check, deduction)
            violation_checker.check_violations()
            if violation_checker.count_violations(-1) == 0:
                style_check = "Alles sieht gut aus -- weiter so!\n"
            violation_string = violation_checker.list_violation()
            violations_checkers.update({folder.name.split('_')[0]: violation_checker})
            style_check += f'\n{violation_string}'
            outfile.write(style_check)


def remove_unnecessary_violations(style_check: str):
    """
    Function to delete all lines with a violation to ignore
    author: Lukas Horst
    """
    lines = style_check.splitlines()

    filtered_lines = []

    skip_count = 0

    e501_pattern = re.compile(r"E501 line too long \((\d+) > 79 characters\)")

    for i, line in enumerate(lines):
        if skip_count > 0:
            skip_count -= 1
            continue
        # Removing lines violations which are shorter than 100
        match = e501_pattern.search(line)
        if match:
            line_length = int(match.group(1))
            if line_length <= 99:
                skip_count = 2
                continue
            else:
                line = line.replace('79', '99')
        # Upper case violations
        elif "C0103" in line and "doesn't conform to UPPER_CASE naming style" in line:
            continue
        # Allowing variable and argument names with only one char
        elif ("C0103" in line and "doesn't conform to snake_case naming style" in line
              and ('Argument name "' in line or 'Variable name "' in line)):
            start_index = line.find('"') + 1
            end_index = line.find('"', start_index)
            argument_name = line[start_index:end_index]
            if len(argument_name) == 1:
                continue
        # Allowing all module names
        elif "C0103" in line and "Module name" in line:
            continue
        # Ignoring a missing whitespace after : in a print command
        elif "E231" in line and "after ':'" in line and "print(" in lines[i + 1]:
            skip_count = 2
            continue
        filtered_lines.append(line)
    return "\n".join(filtered_lines)


def fix_path(path: str) -> str:
    return unicodedata.normalize('NFC', path).replace('U╠ê', 'Ü').replace('u╠ê', 'ü').replace(
        '*', '').replace('"', '')


def safe_extract_zip(zip_obj: zipfile.ZipFile, parent: pathlib.Path):
    parent.mkdir(parents=True, exist_ok=True)
    files = [x for x in zip_obj.infolist() if not x.is_dir()]
    for f in files:
        f_out = parent / pathlib.Path(fix_path(f.filename))
        f_out.parent.mkdir(parents=True, exist_ok=True)
        with zip_obj.open(f) as fin:
            with open(f_out, 'wb') as fout:
                fout.write(fin.read())


def begin_grading(folder: pathlib.Path, ratings_file: pathlib.Path, check_style: bool,
                  author_pairs: bool, deduction: bool):
    print("Extracting downloads...")
    downloads = list(folder.glob('**/*.zip'))
    count = 0
    total = len(downloads)
    for file in downloads:
        count += 1
        print(f" ({str(count).rjust(len(str(total)))}/{total}) Extracting {file.name}")
        with zipfile.ZipFile(file, 'r') as zip_obj:
            # zip_obj.extractall(file.parent / 'abgaben')
            safe_extract_zip(zip_obj, file.parent / 'abgaben')
    print("Extracting archives...")
    archives = list(folder.glob("**/abgaben/**/*.zip"))
    count = 0
    total = len(archives)
    for file in archives:
        count += 1
        print(f" ({str(count).rjust(len(str(total)))}/{total}) Extracting {file.name}")
        with zipfile.ZipFile(file, 'r') as zip_obj:
            # zip_obj.extractall(file.parent)
            safe_extract_zip(zip_obj, file.parent)
    target_folders = [f for f in itertools.chain.from_iterable(
        (group.iterdir() for group in folder.glob('**/abgaben')))
                      if f.is_dir()]
    if check_style:
        print("Running style check...")
        lint_files(target_folders, author_pairs, deduction)
    else:
        print("(Style check skipped.)")
    print("Copying ratings table...")
    count = 0
    sheet = folder.resolve().name
    for f in target_folders:
        count += 1
        target_name = "Bewertung " + sheet + " " + f.name.split('_')[0] + ratings_file.suffix
        shutil.copy(ratings_file, f / target_name)
        if len(violations_checkers) != 0:
            student_name = f.name.split('_')[0]
            file_path = os.path.join(f, target_name)
            if student_name in violations_checkers:
                update_style_deduction(file_path, violations_checkers[student_name], student_name)
        print(f'({count}/{len(target_folders)}) Copy in {f.name}')
    print("Done!")


def finalise_grading(folder: pathlib.Path):
    issues = 0
    print("Copying grades...")
    folders = list(folder.glob("**/abgaben"))
    for f in folders:
        overall_rating_path = ''
        for file_name in os.listdir(f.parent):
            if file_name.startswith('Bewertungen-'):
                overall_rating_path = os.path.join(f.parent, file_name)
                break
        target = f.parent / 'korrekturen'
        target.mkdir()
        count = 0
        handins = [x for x in f.iterdir() if x.name != '.DS_Store']
        for handin in handins:
            count += 1
            this_target = target / handin.name
            this_target.mkdir()
            # copy the stylecheck datas
            if (handin / 'stylecheck.txt').exists():
                shutil.copy(handin / 'stylecheck.txt', this_target)
            # copy the grading datas
            glob = list(handin.glob('Bewertung *'))
            if len(glob) == 1:
                print(f'({count}/{len(handins)}) Copying from {handin.name}')
                shutil.copy(glob[0], this_target)
                # If the overall rating file is given, the points will be written in
                if len(overall_rating_path) != 0:
                    student_name = handin.name.split('_')[0]
                    update_rating(overall_rating_path, glob[0], student_name)
            elif not glob:
                print(f" ! {handin.name}: no grading file")
                issues += 1
            else:
                print(f" ! {handin.name}: too many grading files")
                issues += 1
    if issues:
        print(f"Issues occurred ({issues}), not building final upload file(s).")
        return
    print("Building upload files...")
    folders = list(folder.glob("**/korrekturen"))
    count = 0
    total = len(folders)
    for f in folders:
        count += 1
        print(f" ({str(count).rjust(len(str(total)))}/{total}) Building {f.parent.name}")
        with zipfile.ZipFile(f.parent / (f.parent.name + ".zip"), 'w') as outfile:
            for person in f.iterdir():
                for file in person.iterdir():
                    outfile.write(file, pathlib.PurePath(person.name) / file.name)


def get_points(file_path: str):
    """
    Returns the total points of the given rating table
    author: Lukas Horst
    """
    data = pd.read_excel(file_path, usecols='A, C')
    total_points = 0
    for i in range(len(data)):
        if data.iat[i, 0] == 'Summe':
            break
        value = data.iat[i, 1]
        if pd.notna(value) and type(value) != str:
            total_points += value
    return max(0, total_points)


def update_style_deduction(file_path: str, violation_checker: ViolationChecker, student_name: str):
    """
    Function to update the deduction for style violations in the given rating table
    author: Lukas Horst
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Sheet1']
    # Updating the name
    ws[f'A1'].value = student_name
    rows = ws.iter_rows(min_row=1, max_row=75, min_col=1, max_col=1)
    for i, row in enumerate(rows):
        cell = row[0]
        if cell.value is not None:
            # Updating the deduction for the author variable
            if '__author__' in cell.value:
                ws[f'C{i + 1}'].value = -violation_checker.count_deduction(3)
                ws[f'C{i + 1}'].font = Font(color='FF0000')
            # All deductions except the author variable and docstrings
            elif 'o.g. Fehler' in cell.value:
                deduction = 0
                for j in range(1, 10):
                    if j == 3 or j == 5:
                        continue
                    deduction -= violation_checker.count_deduction(j)
                ws[f'C{i + 1}'].value = deduction
                ws[f'C{i + 1}'].font = Font(color='FF0000')
            # Deduction for docstrings
            elif 'Abzug bei' in cell.value:
                ws[f'C{i + 1}'].value = -violation_checker.count_deduction(5)
                ws[f'C{i + 1}'].font = Font(color='FF0000')
            # Updating the function for the total points
            elif 'Summe' in cell.value:
                ws[f'C{i + 1}'] = f'=MAX(0, SUM(C1:C{i}))'
                break
    wb.save(file_path)
    wb.close()


def read_csv_file(file_path: str):
    """
    Function to read a csv file and returns a list with each row in a dic
    author: Lukas Horst
    """
    with open(file_path, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        rows = []
        for row in reader:
            rows.append(row)
    return rows


def write_csv_file(file_path: str, data: list[dict[str, str]]):
    """
    Function to (over)write a csv file with the given data
    author: Lukas Horst
    """
    with open(file_path, mode='w', newline='', encoding='utf-8') as file:
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)


def update_rating(overall_rating_path: str, student_rating_path: str, student_name: str):
    """
    Function to update the points of the given student
    author: Lukas Horst
    """
    csv_data = read_csv_file(overall_rating_path)
    for row in csv_data:
        if row['Vollständiger Name'] == student_name:
            points = str(get_points(student_rating_path)).replace('.', ',')
            row['Bewertung'] = points
            write_csv_file(overall_rating_path, csv_data)
            return


def main():
    """The function main is where execution begins."""
    print('EPRgrader v3/221031 running on ', datetime.now(), ' [', platform.platform(terse=True),
          ' ',
          platform.machine(), ']', sep='')
    parser = argparse.ArgumentParser(description="Assist in grading EPR assignments.")
    # parser.add_argument('verb', type=str, choices=('begin', 'relint', 'finalise'))
    parser.add_argument('-f', '--folder', type=str,
                        help='the folder in which to operate (default: the current folder)',
                        default='.')
    subparsers = parser.add_subparsers(metavar='verb', dest='verb', required=True)
    begin_parser = subparsers.add_parser('begin', help='begin a new grading process')
    begin_parser.add_argument('--table', metavar='file',
                              help='Ratings table file to copy to each folder',
                              required=True)
    begin_parser.add_argument('--stylecheck', action=argparse.BooleanOptionalAction, default=True,
                              help='whether or not to run style checks')
    begin_parser.add_argument('--pairs', action=argparse.BooleanOptionalAction, default=False,
                              help='whether or not to validate __author__ variables for pairs')
    begin_parser.add_argument('--deduction', action=argparse.BooleanOptionalAction, default=True,
                              help='whether or not to give deduction on the style')
    lint_parser = subparsers.add_parser('relint', help='re-run pylint')
    lint_parser.add_argument('--pairs', action=argparse.BooleanOptionalAction, default=False,
                             help='whether or not to validate __author__ variables for pairs')
    subparsers.add_parser('finalise', help='package results for upload')
    args = parser.parse_args()
    if args.verb == 'begin':
        begin_grading(pathlib.Path(args.folder), pathlib.Path(args.table), args.stylecheck,
                      args.pairs, args.deduction)
    elif args.verb == 'relint':
        lint_files([f for f in itertools.chain.from_iterable(
            (group.iterdir() for group in pathlib.Path(args.folder).glob('**/abgaben'))) if
                    f.is_dir()], args.pairs, args.deduction)
    elif args.verb == 'finalise':
        finalise_grading(pathlib.Path(args.folder))


if __name__ == "__main__":
    main()
